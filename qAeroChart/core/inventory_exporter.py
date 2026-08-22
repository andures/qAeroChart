# -*- coding: utf-8 -*-
"""
Inventory exporters (Issue #110).

Writes layer-inventory rows (see ``core/layer_inventory.py``) to CSV
(stdlib, always available) or XLSX with per-type color fills (openpyxl,
optional dependency — QGIS installs do not ship it by default).
"""
from __future__ import annotations

import csv

from .layer_inventory import HEADERS

try:
    from openpyxl import Workbook  # noqa: F401 - re-exported for write_xlsx internals
    from openpyxl.styles import PatternFill

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Same fills as the contributor's Get_Layers_Path_xls.py.
_TYPE_FILLS = {
    "Point": "FFCCCC",
    "Line": "CCE5FF",
    "Polygon": "CCFFCC",
    "Raster": "FFFACD",
}

_SHEET_NAME = "Layer Info"


def openpyxl_missing_reason() -> str:
    """Human-readable hint shown when the XLSX action is unavailable."""
    return (
        "XLSX export requires the 'openpyxl' package "
        "(install with: pip install openpyxl). CSV export is always available."
    )


def write_csv(rows: list[list[str]], file_path: str) -> str:
    """Write *rows* as UTF-8 CSV (with header) to *file_path*. Returns path."""
    with open(file_path, mode="w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(HEADERS)
        writer.writerows(rows)
    return file_path


def write_xlsx(rows: list[list[str]], file_path: str) -> str:
    """Write *rows* to a formatted single-sheet workbook.

    Raises RuntimeError when openpyxl is not installed — callers must check
    ``HAS_OPENPYXL`` first (the menu action is disabled in that case).
    """
    if not HAS_OPENPYXL:
        raise RuntimeError(openpyxl_missing_reason())

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _SHEET_NAME
    sheet.append(HEADERS)

    fills = {
        name: PatternFill(start_color=hex6, end_color=hex6, fill_type="solid")
        for name, hex6 in _TYPE_FILLS.items()
    }

    for row_data in rows:
        sheet.append(row_data)
        fill = fills.get(row_data[1])
        if fill is not None:
            for col in range(1, len(row_data) + 1):
                sheet.cell(row=sheet.max_row, column=col).fill = fill

    workbook.save(file_path)
    return file_path
