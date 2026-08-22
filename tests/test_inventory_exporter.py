# -*- coding: utf-8 -*-
"""Unit tests for qAeroChart.core.inventory_exporter (Issue #110)."""
from __future__ import annotations

import csv
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest

import tests.mocks.qgis_mock  # noqa: F401

from qAeroChart.core import inventory_exporter as ie
from qAeroChart.core.layer_inventory import HEADERS

ROWS = [
    ["Airports", "Polygon", "memory", "EPSG:3857", "/shp/ap.shp", ""],
    ["VOR", "Point", "ogr", "EPSG:4326", "/data/vor.shp", "Nav"],
]


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------


class TestWriteCsv:
    def test_writes_header_and_rows(self, tmp_path):
        target = tmp_path / "inv.csv"
        result = ie.write_csv(ROWS, str(target))
        assert result == str(target)
        with open(target, encoding="utf-8", newline="") as fh:
            content = list(csv.reader(fh))
        assert content[0] == HEADERS
        assert content[1] == ROWS[0]
        assert content[2] == ROWS[1]

    def test_utf8_content_round_trip(self, tmp_path):
        rows = [["Aeródromo", "Point", "ogr", "EPSG:4326", "/d/á.shp", ""]]
        target = tmp_path / "utf.csv"
        ie.write_csv(rows, str(target))
        with open(target, encoding="utf-8", newline="") as fh:
            back = list(csv.reader(fh))
        assert back[1] == rows[0]

    def test_empty_rows_still_write_header(self, tmp_path):
        target = tmp_path / "empty.csv"
        ie.write_csv([], str(target))
        with open(target, encoding="utf-8", newline="") as fh:
            assert next(csv.reader(fh)) == HEADERS


# ---------------------------------------------------------------------------
# XLSX
# ---------------------------------------------------------------------------


class TestWriteXlsx:
    def test_raises_when_openpyxl_missing(self, monkeypatch):
        monkeypatch.setattr(ie, "HAS_OPENPYXL", False)
        with pytest.raises(RuntimeError, match="openpyxl"):
            ie.write_xlsx(ROWS, "/tmp/never.xlsx")

    @pytest.mark.skipif(
        not ie.HAS_OPENPYXL, reason="openpyxl not installed in this environment"
    )
    def test_workbook_contents_and_fills(self, tmp_path):
        from openpyxl import load_workbook

        target = tmp_path / "inv.xlsx"
        ie.write_xlsx(ROWS, str(target))

        wb = load_workbook(str(target))
        ws = wb.active
        assert ws.title == "Layer Info"
        assert [c.value for c in ws[1]] == HEADERS
        assert ws.cell(row=2, column=1).value == "Airports"

        fill = ws.cell(row=2, column=1).fill
        assert fill.fill_type == "solid"
        assert fill.start_color.rgb.endswith("CCFFCC")  # Polygon green
        point_fill = ws.cell(row=3, column=1).fill
        assert point_fill.start_color.rgb.endswith("FFCCCC")  # Point pink

    @pytest.mark.skipif(
        ie.HAS_OPENPYXL, reason="only meaningful without openpyxl"
    )
    def test_flag_false_in_clean_env(self):
        assert ie.HAS_OPENPYXL is False


class TestOpenpyxlMissingReason:
    def test_message_mentions_pip_install(self):
        assert "pip install openpyxl" in ie.openpyxl_missing_reason()
